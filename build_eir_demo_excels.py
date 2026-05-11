#!/usr/bin/env python3
"""
build_eir_demo_excels.py — Build two PortF cashflow Excel files for the
EIR Calculation Trace demo:

  1. portf-cashflow-aurora-multi-tranche.xlsx
     £120m multi-tranche infrastructure loan
     - Senior £80m fixed 5.75% + £400k EIR-included arrangement fee
     - Mezz   £40m fixed 9.25% + £600k EIR-included arrangement fee
     - 5-year bullet · 30/360
     - Loads against the auroraMultiTranche seed (matched by sheet name)

  2. portf-cashflow-helios-rfr.xlsx
     £30m bilateral RFR loan, SONIA + ratcheted margin
     - Margin ratchet: 275bps → 300bps → 325bps
     - 5-year bullet · ACT/365
     - Loads against the heliosBridge seed (matched by sheet name)

Follows the canonical PortF Excel template described in
PORTF-EXCEL-TEMPLATE-GUIDE.md.

Run:
  pip3 install openpyxl --break-system-packages
  python3 build_eir_demo_excels.py
"""

import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Where to write the files (override with env var when running in sandbox)
OUT_DIR = os.environ.get(
    'EIR_DEMO_OUT',
    '/Users/ferhatansari/Claude Cowork folder/Claude Income Calculator'
)

# ─── STYLES ──────────────────────────────────────────────────────────────────
LABEL = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
LABEL_FILL = PatternFill('solid', fgColor='21295C')
SETUP_LBL = Font(name='Calibri', size=10, bold=True, color='1F2A44')
SETUP_VAL = Font(name='Calibri', size=10, color='1A1F36')
FEE_HEADER = Font(name='Calibri', size=10, bold=True, color='065A82')
FEE_HEADER_FILL = PatternFill('solid', fgColor='EEF6FA')
DATA_HEADER = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
DATA_HEADER_FILL = PatternFill('solid', fgColor='1C7293')
DATA_CELL = Font(name='Calibri', size=9, color='1A1F36')
THIN = Side(border_style='thin', color='D4DAE3')
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
LEFT = Alignment(horizontal='left',  vertical='center', wrap_text=False)
CTR  = Alignment(horizontal='center',vertical='center')
RT   = Alignment(horizontal='right', vertical='center')

DATE_FMT = 'yyyy-mm-dd'
NUM_FMT = '#,##0.00'
INT_FMT = '#,##0'
PCT_FMT = '0.0000%'

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def write_setup(ws, rows):
    """Section A: setup metadata in columns A/B."""
    for i, (label, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=label).font = SETUP_LBL
        ws.cell(row=i, column=1).alignment = LEFT
        c = ws.cell(row=i, column=2, value=value)
        c.font = SETUP_VAL
        c.alignment = LEFT
        if isinstance(value, (date,)):
            c.number_format = DATE_FMT
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            c.number_format = NUM_FMT if value > 1000 else NUM_FMT

def write_fee_block(ws, col, fee):
    """Section B: 9-row fee setup block in one column (rows 1..9).
    `fee` is a dict with keys: name, postingType, basis, structure, rate,
                              accrualFreq, settleFreq, settleType, firstSettleDate
    """
    cells = [
        (1, fee['name'],            FEE_HEADER, FEE_HEADER_FILL),
        (2, fee['postingType'],     SETUP_VAL,  None),
        (3, fee['basis'],           SETUP_VAL,  None),
        (4, fee['structure'],       SETUP_VAL,  None),
        (5, fee['rate'],            SETUP_VAL,  None),
        (6, fee['accrualFreq'],     SETUP_VAL,  None),
        (7, fee['settleFreq'],      SETUP_VAL,  None),
        (8, fee['settleType'],      SETUP_VAL,  None),
        (9, fee['firstSettleDate'], SETUP_VAL,  None),
    ]
    for row, value, font, fill in cells:
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        c.alignment = LEFT
        if fill: c.fill = fill
        if row == 5:                                # Rate row
            c.number_format = PCT_FMT
        elif row == 9 and isinstance(value, date):  # First Settle Date row
            c.number_format = DATE_FMT

def write_data_header(ws, header_row, fee_columns):
    """Section D: data-table header row (row 17 by default)."""
    headers = [
        (1, 'Date'),
        (2, 'Initial Purchase'),
        (3, 'Drawdown'),
        (4, 'Principal Payment'),
        (5, 'Day count'),
        (6, 'Amount'),    # Principal Balance
        (7, 'Amount'),    # Unfunded Balance
    ]
    for col, txt in headers:
        c = ws.cell(row=header_row, column=col, value=txt)
        c.font = DATA_HEADER
        c.fill = DATA_HEADER_FILL
        c.alignment = CTR
        c.border = BORDER
    # Section labels at row 11-12 (optional but matches Volt example)
    ws.cell(row=11, column=6, value='Balance Basis').font = SETUP_LBL
    ws.cell(row=11, column=6).alignment = LEFT
    ws.cell(row=12, column=6, value='Principal Balance').font = SETUP_VAL
    ws.cell(row=12, column=6).alignment = LEFT
    ws.cell(row=11, column=7, value='Balance Basis').font = SETUP_LBL
    ws.cell(row=11, column=7).alignment = LEFT
    ws.cell(row=12, column=7, value='Unfunded Balance').font = SETUP_VAL
    ws.cell(row=12, column=7).alignment = LEFT
    # Per-fee Rate / Amount column pairs (H/I, J/K, L/M, N/O, …)
    for col, label in fee_columns:
        c1 = ws.cell(row=header_row, column=col,     value='Rate')
        c2 = ws.cell(row=header_row, column=col + 1, value='Amount')
        for c in (c1, c2):
            c.font = DATA_HEADER
            c.fill = DATA_HEADER_FILL
            c.alignment = CTR
            c.border = BORDER
        # Section label above
        ws.cell(row=11, column=col, value=label).font = SETUP_LBL
        ws.cell(row=11, column=col).alignment = LEFT

def write_data_row(ws, row, fields):
    """Write one data-table row. `fields` is dict keyed by column index."""
    for col, value in fields.items():
        c = ws.cell(row=row, column=col, value=value)
        c.font = DATA_CELL
        c.alignment = RT
        c.border = BORDER
        if col == 1 and isinstance(value, date):
            c.number_format = DATE_FMT
            c.alignment = CTR
        elif col == 5:
            c.number_format = INT_FMT
        elif col in (2, 3, 4, 6, 7):
            c.number_format = NUM_FMT
        elif col % 2 == 0:    # even cols H/J/L/N = Rate
            c.number_format = NUM_FMT
        else:
            c.number_format = NUM_FMT

def set_widths(ws, widths):
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

# ─── 1. AURORA RENEWABLES (multi-tranche) ────────────────────────────────────
def build_aurora():
    wb = Workbook()
    # Sheet name MUST match the seed instrument's id for the parser to match
    ws = wb.active
    ws.title = 'auroraMultiTranche'

    # Section A · Setup metadata (rows 1-15, cols A/B)
    write_setup(ws, [
        ('Company',              'Aurora Renewables Phase 1'),
        ('Debt Type',            'Multi-Tranche'),
        ('Loan Start Date',      date(2026, 3, 1)),
        ('Loan End Date',        date(2031, 3, 1)),
        ('Day Count Convention', '30/360'),
        ('Interest Accrues',     'Same day'),
        ('Total Commitment',     120_000_000),
        ('Currency',             'GBP'),
        ('Legal Entity',         'NWF Sustainable Infrastructure'),
        ('LEID',                 42),
        ('Position',             'NWF 100% Bilateral Position · Aurora Renewables'),
        ('Position ID',          'POS-NWF-AURORA-MT'),
        ('Income Security',      'Aurora Renewables Multi-Tranche Facility (£80m Senior 5.75% + £40m Mezz 9.25%)'),
        ('Security ID',          'SEC-AURORA-RENEW-MT'),
        ('Counterparty',         'Aurora Renewables Holdings PLC'),
    ])

    # Section B · Fee blocks (rows 1-9, cols H,J,L,N — 2 cols per fee)
    write_fee_block(ws, 8,  {       # col H — Senior interest
        'name':            'Senior Coupon (Underlying Debt)',
        'postingType':     'Interest',
        'basis':           'Principal Balance',
        'structure':       'Fixed',
        'rate':            0.0575,
        'accrualFreq':     'Quarterly',
        'settleFreq':      'Quarterly',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 6, 30),
    })
    write_fee_block(ws, 10, {       # col J — Mezz interest
        'name':            'Mezz Coupon (Underlying Debt 2)',
        'postingType':     'Interest',
        'basis':           'Principal Balance',
        'structure':       'Fixed',
        'rate':            0.0925,
        'accrualFreq':     'Quarterly',
        'settleFreq':      'Quarterly',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 6, 30),
    })
    write_fee_block(ws, 12, {       # col L — Senior arrangement fee
        'name':            'Senior Arrangement Fee',
        'postingType':     'Fee',
        'basis':           'Commitment Balance',
        'structure':       'Fixed',
        'rate':            0.0050,           # 0.5% × £80m = £400k
        'accrualFreq':     'One-off',
        'settleFreq':      'One-off',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 3, 1),
    })
    write_fee_block(ws, 14, {       # col N — Mezz arrangement fee
        'name':            'Mezz Arrangement Fee',
        'postingType':     'Fee',
        'basis':           'Commitment Balance',
        'structure':       'Fixed',
        'rate':            0.0150,           # 1.5% × £40m = £600k
        'accrualFreq':     'One-off',
        'settleFreq':      'One-off',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 3, 1),
    })

    # Section D · Data table (header row 17, data rows 18+)
    fee_cols = [(8, 'Senior Coupon'), (10, 'Mezz Coupon'),
                (12, 'Senior Arr.'),  (14, 'Mezz Arr.')]
    write_data_header(ws, 17, fee_cols)

    # Build quarterly periods 2026-03-01 → 2031-03-01
    period_ends = [
        date(2026, 6, 30), date(2026, 9, 30), date(2026,12,31),
        date(2027, 3,31), date(2027, 6,30), date(2027, 9,30), date(2027,12,31),
        date(2028, 3,31), date(2028, 6,30), date(2028, 9,30), date(2028,12,31),
        date(2029, 3,31), date(2029, 6,30), date(2029, 9,30), date(2029,12,31),
        date(2030, 3,31), date(2030, 6,30), date(2030, 9,30), date(2030,12,31),
        date(2031, 3, 1),
    ]
    SENIOR = 80_000_000
    MEZZ   = 40_000_000
    SENIOR_RATE = 0.0575
    MEZZ_RATE   = 0.0925

    # Row 18: Day 1 drawdown — entire £120m drawn at signing, arrangement fees paid
    write_data_row(ws, 18, {
        1: date(2026, 3, 1),
        2: 0,
        3: -120_000_000,                    # PortF convention: negative = draw
        4: 0,
        5: 0,
        6: 120_000_000,                     # Principal balance after draw
        7: 0,                               # Unfunded
        # Senior + Mezz interest = 0 on day 1
         8: 0.0575,  9: 0,
        10: 0.0925, 11: 0,
        # Arrangement fees paid upfront
        12: 0.0050, 13: 400_000,            # Senior Arrangement: £400k
        14: 0.0150, 15: 600_000,            # Mezz Arrangement:   £600k
    })

    # Subsequent quarterly rows: interest accrual on £80m senior + £40m mezz
    prev = date(2026, 3, 1)
    for i, eop in enumerate(period_ends):
        # 30/360 day count for each period
        days_30_360 = ((eop.year - prev.year)*360 +
                       (eop.month - prev.month)*30 +
                       (eop.day - prev.day))
        days_30_360 = max(0, days_30_360)
        senior_int = SENIOR * SENIOR_RATE * days_30_360 / 360
        mezz_int   = MEZZ   * MEZZ_RATE   * days_30_360 / 360
        is_last = (i == len(period_ends) - 1)
        write_data_row(ws, 19 + i, {
            1: eop,
            2: 0,
            3: 0,
            4: 120_000_000 if is_last else 0,    # Principal repaid at maturity
            5: days_30_360,
            6: 0 if is_last else 120_000_000,
            7: 0,
             8: SENIOR_RATE,  9: round(senior_int, 2),
            10: MEZZ_RATE,   11: round(mezz_int, 2),
            12: 0, 13: 0,
            14: 0, 15: 0,
        })
        prev = eop

    set_widths(ws, {1:13, 2:14, 3:16, 4:18, 5:10, 6:16, 7:14,
                    8:9, 9:14, 10:9, 11:14, 12:9, 13:14, 14:9, 15:14})
    return wb


# ─── 2. HELIOS SOLAR BRIDGE (RFR floating) ───────────────────────────────────
def build_helios():
    wb = Workbook()
    ws = wb.active
    ws.title = 'heliosBridge'

    # Section A · Setup metadata
    write_setup(ws, [
        ('Company',              'Helios Solar Bridge'),
        ('Debt Type',            'Loan'),
        ('Loan Start Date',      date(2026, 4, 15)),
        ('Loan End Date',        date(2031, 4, 15)),
        ('Day Count Convention', 'ACT/365'),
        ('Interest Accrues',     'Compounded'),
        ('Total Commitment',     30_000_000),
        ('Currency',             'GBP'),
        ('Legal Entity',         'NWF Sustainable Infrastructure'),
        ('LEID',                 42),
        ('Position',             'NWF 100% Bilateral Position · Helios Solar Bridge'),
        ('Position ID',          'POS-NWF-HELIOS-RFR'),
        ('Income Security',      'Helios Solar Bridge Facility (£30m, Compounded SONIA + Ratcheted Margin)'),
        ('Security ID',          'SEC-HELIOS-SOLAR-BRIDGE'),
        ('Counterparty',         'Barclays Bank'),
    ])

    # Section B · Fee blocks
    # The interest line is the underlying SONIA + margin coupon. Margin ratchets
    # are captured in the data table's Rate column period-by-period.
    write_fee_block(ws, 8, {        # col H — Underlying Debt (SONIA + margin)
        'name':            'Underlying Debt (Compounded SONIA + Margin)',
        'postingType':     'Interest',
        'basis':           'Principal Balance',
        'structure':       'Floating (SONIA + Spread)',
        'rate':            0.0750,           # Initial all-in: 4.75% SONIA + 2.75% margin = 7.50%
        'accrualFreq':     'Daily',
        'settleFreq':      'Quarterly',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 7, 31),
    })
    write_fee_block(ws, 10, {       # col J — Commitment fee on undrawn (year 1 only)
        'name':            'Commitment Fee',
        'postingType':     'Non-use Fee',
        'basis':           'Unfunded Balance',
        'structure':       'Floating (SONIA + Spread)',
        'rate':            0.0010,           # 35% × ~2.75% margin ≈ 0.1%
        'accrualFreq':     'Quarterly',
        'settleFreq':      'Quarterly',
        'settleType':      'Cash',
        'firstSettleDate': date(2026, 7, 31),
    })

    # Section D · Data table
    fee_cols = [(8, 'Underlying Debt'), (10, 'Commitment Fee')]
    write_data_header(ws, 17, fee_cols)

    PRINCIPAL = 30_000_000
    SONIA = 0.0475                 # illustrative current SONIA reset (4.75%)

    # Margin ratchet schedule (matches the seed instrument's marginSchedule[])
    def margin_for(d):
        if d < date(2028, 4, 15):  return 0.0275
        if d < date(2030, 4, 15):  return 0.0300
        return 0.0325

    period_ends = [
        date(2026, 4,15),   # signing / drawdown
        date(2026, 7,31), date(2026,10,31), date(2027, 1,31), date(2027, 4,30),
        date(2027, 7,31), date(2027,10,31), date(2028, 1,31), date(2028, 4,15),  # ratchet boundary
        date(2028, 7,31), date(2028,10,31), date(2029, 1,31), date(2029, 4,30),
        date(2029, 7,31), date(2029,10,31), date(2030, 1,31), date(2030, 4,15),  # ratchet boundary
        date(2030, 7,31), date(2030,10,31), date(2031, 1,31),
        date(2031, 4,15),   # final maturity
    ]

    # Row 18: Day 1 drawdown
    write_data_row(ws, 18, {
        1: date(2026, 4, 15),
        2: 0,
        3: -30_000_000,
        4: 0,
        5: 0,
        6: 30_000_000,
        7: 0,
        8: 0.0750,  9: 0,
       10: 0.0010, 11: 0,
    })

    prev = date(2026, 4, 15)
    for i, eop in enumerate(period_ends):
        if eop == prev:
            continue
        days_act = (eop - prev).days
        margin = margin_for(prev)
        all_in = SONIA + margin
        int_amt = PRINCIPAL * all_in * days_act / 365
        commit_rate = margin * 0.35     # 35% of margin
        is_last = (eop == date(2031, 4, 15))
        write_data_row(ws, 18 + i, {
            1: eop,
            2: 0,
            3: 0,
            4: 30_000_000 if is_last else 0,
            5: days_act,
            6: 0 if is_last else 30_000_000,
            7: 0,
             8: round(all_in, 6),
             9: round(int_amt, 2),
            10: round(commit_rate, 6),
            11: 0,                            # no commitment fee post day-1 (fully drawn)
        })
        prev = eop

    set_widths(ws, {1:13, 2:14, 3:16, 4:18, 5:10, 6:16, 7:14,
                    8:9, 9:14, 10:9, 11:14})
    return wb


# ─── BUILD ───────────────────────────────────────────────────────────────────
def main():
    aurora_path = os.path.join(OUT_DIR, 'portf-cashflow-aurora-multi-tranche.xlsx')
    helios_path = os.path.join(OUT_DIR, 'portf-cashflow-helios-rfr.xlsx')

    build_aurora().save(aurora_path)
    print(f'  ✓ Wrote {aurora_path}')

    build_helios().save(helios_path)
    print(f'  ✓ Wrote {helios_path}')


if __name__ == '__main__':
    main()
